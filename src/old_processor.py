import pandas as pd
import re
import os
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

def clean_ecn_data(input_path, output_path):
    """Limpia los datos del CSV crudo extraído de Outlook."""
    df = pd.read_csv(input_path)
    
    # Regex para ECN y Job (Versión ultra-flexible)
    df['ECN_Number'] = df['Subject'].str.extract(r'ECN\s*#?\s*([\d-]+)', flags=re.IGNORECASE)
    df['Job_Number'] = df['Subject'].str.extract(r'Job\s+([A-Z0-9-/]+)', flags=re.IGNORECASE)
    
    # Cambia la parte de las fechas por esta versión más "amigable" con Excel:
    df['Received_Date'] = pd.to_datetime(df['Received_Date'], errors='coerce').dt.tz_localize(None)
    df['Month'] = df['Received_Date'].dt.strftime('%Y-%m')
    
    # KPI: Días transcurridos (ahora sin problemas de zona horaria)
    now = pd.Timestamp.now()
    df['Days_Open'] = (now - df['Received_Date']).dt.days
        
    # Limpieza de nulos y duplicados en el set nuevo
    df_clean = df.dropna(subset=['ECN_Number']).copy()
    df_clean = df_clean.sort_values('Received_Date', ascending=False).drop_duplicates('ECN_Number')
    
    # Guardamos el CSV procesado (como respaldo técnico)
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    df_clean.to_csv(output_path, index=False)
    
    return df_clean

def update_excel_tracker(df_new, excel_path):
    """Actualiza el Excel 'inyectando' datos en la tabla oficial sin borrar formatos."""
    sheet_name = "ECNs"
    table_name = "Tbl_ECN_Master"

    if not os.path.exists(excel_path):
        # Si no existe, lo creamos por primera vez
        df_new['Classification'] = "New BOM Release"
        df_new['Urgency'] = "1_Green"
        df_new['Actions_Required'] = "Regenerate_PKs"
        df_new['Status'] = "Pending"
        df_new.to_excel(excel_path, index=False, sheet_name=sheet_name)
        print(f"-> Archivo Maestro creado desde cero.")
        return

    # --- PROCESO DE INYECCIÓN CON OPENPYXL ---
    # 1. Cargar el libro y manejar el nombre de la hoja
    wb = load_workbook(excel_path)
    if sheet_name not in wb.sheetnames:
        ws = wb.active
        ws.title = sheet_name # Cambiamos el nombre a 'ECNs'
    else:
        ws = wb[sheet_name]

    # 2. Leer datos actuales para evitar duplicados (usando pandas)
    df_existing = pd.read_excel(excel_path, sheet_name=sheet_name)
    new_records = df_new[~df_new['ECN_Number'].isin(df_existing['ECN_Number'])].copy()

    if not new_records.empty:
        new_records['Classification'] = "New BOM Released"
        new_records['Urgency'] = "1_Green"
        new_records['Actions_Required'] = "Regenerate_PKs"
        new_records['Status'] = "Pending"

        # 3. Encontrar la última fila con datos
        last_row = ws.max_row

        # 4. Escribir los registros nuevos fila por fila al final
        for r_idx, row in enumerate(new_records.values, start=last_row + 1):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        # 5. Actualizar el rango de la Tabla "Tbl_ECN_Master"
        # Esto es vital para que Power BI vea los nuevos datos
        if table_name in ws.tables:
            tab = ws.tables[table_name]
            # Calculamos el nuevo rango (de A1 hasta la última columna/fila)
            last_col_letter = ws.cell(row=1, column=ws.max_column).column_letter
            new_range = f"A1:{last_col_letter}{ws.max_row}"
            tab.ref = new_range
        
        wb.save(excel_path)
        print(f"-> ¡Éxito! Se inyectaron {len(new_records)} ECNs y se actualizó la tabla.")
    else:
        print("-> No hay ECNs nuevos. El archivo se mantuvo intacto.")