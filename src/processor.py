import pandas as pd
import os
import re
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils.dataframe import dataframe_to_rows

def clean_ecn_data(input_path, output_path):
    """Limpia los datos del CSV crudo extraído de Outlook."""
    df = pd.read_csv(input_path)
    
    # Regex para ECN y Job (Versión ultra-flexible)
    df['ECN_Number'] = df['Subject'].str.extract(r'ECN\s*#?\s*([\d-]+)', flags=re.IGNORECASE)
    df['Job_Number'] = df['Subject'].str.extract(r'Job\s+([A-Z0-9-/]+)', flags=re.IGNORECASE)
    
    # Cambia la parte de las fechas por esta versión más "amigable" con Excel:
    df['Received_Date'] = pd.to_datetime(df['Received_Date'], errors='coerce').dt.tz_localize(None)
    df['Month'] = df['Received_Date'].dt.strftime('%Y-%m')
    
    # KPI: Días transcurridos (ahora sin problemas de zona horaria)
    now = pd.Timestamp.now()
    df['Days_Open'] = (now - df['Received_Date']).dt.days
        
    # Limpieza de nulos y duplicados en el set nuevo
    df_clean = df.dropna(subset=['ECN_Number']).copy()
    df_clean = df_clean.sort_values('Received_Date', ascending=False).drop_duplicates('ECN_Number')
    
    # Guardamos el CSV procesado (como respaldo técnico)
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    df_clean.to_csv(output_path, index=False)
    
    return df_clean

def update_excel_tracker(input_path, output_path):
    """Actualiza o crea el Excel asegurando la existencia de la Tabla oficial."""
    sheet_name = "ECNs"
    table_name = "Tbl_ECN_Master"

    # Preparar los datos nuevos (si el archivo no existe, estos serán los primeros)
    if not os.path.exists(output_path):
        df_to_add = input_path.copy()
        df_to_add['Classification'] = "BOM_Rev"
        df_to_add['Urgency'] = "1_Green"
        df_to_add['Actions_Required'] = "Regenerate_PKs"
        df_to_add['Status'] = "Pending"
        is_new_file = True
    else:
        df_existing = pd.read_excel(output_path, sheet_name=sheet_name)
        df_to_add = input_path[~input_path['ECN_Number'].isin(df_existing['ECN_Number'])].copy()
        if not df_to_add.empty:
            df_to_add['Classification'] = "BOM_Rev"
            df_to_add['Urgency'] = "1_Green"
            df_to_add['Actions_Required'] = "Regenerate_PKs"
            df_to_add['Status'] = "Pending"
        is_new_file = False

    if df_to_add.empty and not is_new_file:
        print("-> No hay ECNs nuevos. El archivo se mantuvo intacto.")
        return

    # --- MANEJO CON OPENPYXL ---
    if is_new_file:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        # Escribir encabezados y datos
        for r in dataframe_to_rows(df_to_add, index=False, header=True):
            ws.append(r)
    else:
        wb = load_workbook(output_path)
        ws = wb[sheet_name]
        # Escribir solo los datos nuevos (sin encabezados)
        for r in dataframe_to_rows(df_to_add, index=False, header=False):
            ws.append(r)

    # ACTUALIZAR O CREAR LA TABLA (El corazón de la automatización)
    last_col_letter = ws.cell(row=1, column=ws.max_column).column_letter
    new_range = f"A1:{last_col_letter}{ws.max_row}"

    if table_name in ws.tables:
        # Si la tabla ya existe, solo actualizamos su rango
        ws.tables[table_name].ref = new_range
    else:
        # Si es archivo nuevo, creamos la tabla desde cero
        tab = Table(displayName=table_name, ref=new_range)
        # Agregamos un estilo visual por defecto (puedes cambiarlo)
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        tab.tableStyleInfo = style
        ws.add_table(tab)

    wb.save(output_path)
    print(f"-> Proceso completado: Archivo {'creado' if is_new_file else 'actualizado'} y tabla '{table_name}' lista.")

if __name__ == "__main__":
    # Rutas de archivos en carpeta compartida de Planeación
    # raw_csv = r"O:\11-SFM_Level_2_Planning\ECN_Project\data\raw\ecn_raw_data.csv"
    # processed_csv = r"O:\11-SFM_Level_2_Planning\ECN_Project\data\processed\ecn_cleaned_data.csv"
    # master_excel = r"O:\11-SFM_Level_2_Planning\ECN_Project\ECN_Tracker_Master.xlsx"
    
    # Rutas de archivos en local computadora de aorduna
    raw_csv = r"C:\Users\10147115\ECN_Project\data\raw\ecn_raw_data.csv"
    processed_csv = r"C:\Users\10147115\ECN_Project\data\processed\ecn_cleaned_data.csv"
    master_excel = r"C:\Users\10147115\ECN_Project\ECN_Tracker_Master.xlsx"
    
    # Ejecución del Pipeline
    print("Iniciando procesamiento de datos...")
    df_cleaned = clean_ecn_data(raw_csv, processed_csv)
    
    print("Actualizando Tracker Maestro de Excel...")
    update_excel_tracker(df_cleaned, master_excel)
    
    print("\nProceso finalizado correctamente.")